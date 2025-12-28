#!python
"""
BENİ OKU
https://chromedriver.chromium.org/downloads #iptal kurulu chrome kullanıyor.

pyinstaller kurulu değil ise kur. : pip install pyinstaller
pyinstaller -F sok.py  #ile .exe'ye çevir.
pyinstaller --name=sok1000 -F -i favicon.ico sok.py

https://chromedriver.storage.googleapis.com/index.html?path=114.0.5735.90/ #chromeDriver indirme linki
"""

import sys
sys.path
sys.executable

import os  # Dosya işlemleri için kullanılır.
import datetime  # Tarih işlemleri için kullanılır.
#import requests  # HTTP istekleri yapmak için kullanılır. #Kaydırmalı sayfa olduğu için çalışmadı selenium kullanıldı.
from bs4 import BeautifulSoup  # HTML ve XML dosyalarının ayrıştırılması için kullanılır.
from selenium import webdriver  # Web tarayıcıları ile otomatik işlemler yapmak için kullanılır.
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import SessionNotCreatedException
import time  # Zaman işlemleri için kullanılır.
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#Windows uygulamalarda print çalışmadığı için print/sys.stdout.write kullanmak için yaz fonksiyonu yazıldı.
def yaz(metin):
    sys.stdout.write(metin+'\n')

#Url exe'den sonra zaten yazıldı ise url isteme
def get_url():
    if len(sys.argv) > 1:
        return sys.argv[1]
    else:
        return input("Lütfen URL'yi girin veya Enter'a basarak varsayılan adres ile devam edin: ")

#tarayıcı açmak için fonksiyon
def tarayici_ac():
    """
    Chrome tarayıcısını açar ve bilgisayarda yüklü olan Chrome tarayıcısını kullanır.
    """
    print("Tarayıcı açılıyor...")
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    # Bilgisayarda yüklü olan Chrome tarayıcısını kullanır
    driver = webdriver.Chrome(options=options)

    driver.set_window_size(1360,960)
    return driver

def sayfayi_ziyaret_et(driver, url):
    """
    Bu fonksiyon, verilen tarayıcı nesnesiyle verilen URL'yi ziyaret eder ve sayfa başlığını döndürür.
    """
    yaz (url+" sayfa ziyaret ediliyor...")
    driver.get(url) # URL'yi ziyaret eder.
    time.sleep(5) # 5 saniye bekler.
    return driver.title

def sayfayi_kaydir(driver, kez, saniye):
    """
    Bu fonksiyon, verilen tarayıcı nesnesiyle verilen kez kadar sayfayı aşağı kaydırır ve her kaydırma işleminden sonra verilen saniye kadar bekler.
    """
    for i in range(kez): # Verilen kez kadar döngü çalışır.
        driver.execute_script("window.scrollBy(0, 500);") # Sayfayı 500 piksel aşağı kaydırır.
        time.sleep(saniye) # Verilen saniye kadar bekler.
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'CProductCard-module_productCardWrapper__okAmT'))
            )
        except:
            yaz("Sayfa kaydırma sırasında hata oluştu.")
            break

def verileri_al(soup):
    """
    Bu fonksiyon, verilen BeautifulSoup nesnesi üzerinden verileri alır ve veri listesi olarak döndürür.
    """
    yaz ("Veriler çekilmeye başlandı...")
    product_cards = soup.find_all(class_='CProductCard-module_productCardWrapper__okAmT') # Ürün kartlarını bulur.
    
    if not product_cards:
        yaz("Hata: Ürün kartları bulunamadı.")
        return []

    veri_listesi = [] # Boş veri listesi oluşturur.
    for card in product_cards: # Her bir ürün kartı için döngü çalışır.
        urun_ismi = card.find(class_='CProductCard-module_title__u8bMW')  # Ürün ismini bulur.
        fiyat = card.find(class_='CPriceBox-module_price__bYk-c')  # Fiyatı bulur.
        indirimli_fiyat = card.find(class_='CPriceBox-module_discountedPrice__15Ffw')  # İndirimli fiyatı bulur.

        if urun_ismi is not None:  # Eğer 'urun_ismi' elementi varsa:
            urun_ismi = urun_ismi.text.strip()  # Elementin metnini urun_ismi değişkenine atar ve boşlukları siler.
        else:
            urun_ismi = "Bilinmiyor"

        if indirimli_fiyat is not None:  # Eğer indirimli fiyat varsa:
            yeni_fiyat = indirimli_fiyat.text.replace('₺', '').strip()
            eski_fiyat = fiyat.text.replace('₺', '').strip() if fiyat is not None else "0"
        else:
            yeni_fiyat = fiyat.text.replace('₺', '').strip() if fiyat is not None else "0"
            eski_fiyat = "0"

        veri_listesi.append((urun_ismi, yeni_fiyat, eski_fiyat))  # Veri listesine tablo olarak ekler.
    return veri_listesi  # Veri listesini döndürür.


def verileri_yaz(veri_listesi):
    """
    Bu fonksiyon, verilen veri listesini mevcut tarih bazında dosyaya yazar. Örneğin: veriler20230102.txt
    """
    yaz ("Veriler txt dosyasına yazılıyor...")
    # Mevcut tarihi al
    now = datetime.datetime.now()
    dosya_adi = f"veriler{now.year}{now.month:02d}{now.day:02d}.txt" # Mevcut tarih bazında dosya adı oluşturur. Örneğin: veriler20230102.txt

    if os.path.exists(dosya_adi): # Eğer dosya zaten mevcutsa:
        os.remove(dosya_adi) # Dosyayı siler.

    dosya = open(dosya_adi, "w", encoding="utf-8")  # Dosyayı açar ve utf-8 kodlaması ile yazar.
    for veri in veri_listesi: # Her bir veri için döngü çalışır.
        dosya.write(str(veri) + "\n")  # Veriyi dosyaya yazar ve alt satıra geçer.


def excele_yaz_hercalistinda1excel(veri_listesi):
    """
    Bu fonksiyon, verilen veri listesini mevcut tarih bazında Excel dosyasına yazar. Örneğin: veriler20230102.xlsx
    """
    print("Veriler Excel dosyasına yazılıyor...")
    # openpyxl kütüphanesini import et
    from openpyxl import Workbook

    # Excel dosyası oluştur
    workbook = Workbook()

    # Aktif sayfayı al
    sayfa = workbook.active

    # Sayfaya bir sütun adı ekle
    sayfa["A1"] = "Ürün"
    sayfa["B1"] = "Yeni Fiyat"
    sayfa["C1"] = "Eski Fiyat"

    # Veri listesini gez
    for veri in veri_listesi:
        # Veriyi Excel sayfasına ekle
        sayfa.append(veri)

    # Mevcut tarihi al
    now = datetime.datetime.now()
    dosya_adi = f"veriler{now.year}{now.month:02d}{now.day:02d}.xlsx" # Mevcut tarih bazında dosya adı oluşturur. Örneğin: veriler20230102.xlsx

    # Excel dosyasını kaydet
    workbook.save(dosya_adi)


def excele_yaz(veri_listesi, sayfa_basligi):
    """
    Bu fonksiyon, verilen veri listesini mevcut tarih bazında Excel dosyasına yazar. Örneğin: 20241225-1024.xlsx
    """
    print("Veriler Excel dosyasına yazılıyor...")
    # openpyxl kütüphanesini import et
    from openpyxl import load_workbook
    from openpyxl import Workbook

    # Mevcut tarihi al
    now = datetime.datetime.now()
    dosya_adi = f"{now.year}-{now.month:02d}-{now.day:02d}-{now.hour:02d}.xlsx" # Mevcut tarih ve saat bazında dosya adı oluşturur.

    # Excel dosyasını aç veya oluştur
    try:
        workbook = load_workbook(dosya_adi)
    except FileNotFoundError:
        workbook = Workbook()

    # Sayfa başlığının ilk 30 karakterini al
    sayfa_basligi = sayfa_basligi[:30]

    # Yeni sayfa oluştur
    sayfa = workbook.create_sheet(title=sayfa_basligi)
    sayfa["A1"] = "Ürün"
    sayfa["B1"] = "Yeni Fiyat"
    sayfa["C1"] = "Eski Fiyat"
    sayfa.column_dimensions["A"].width = 60 # A1 sütununu genişlet

    # Veri listesini gez
    for veri in veri_listesi:
        # Veriyi Excel sayfasına ekle
        sayfa.append(veri)

    # Excel dosyasını kaydet
    workbook.save(dosya_adi)




yaz ("Bunun gibi bir url yazınız: https://www.sokmarket.com.tr/meyve-ve-sebze-c-20")
url = get_url()

driver = tarayici_ac() # Tarayıcıyı açar.
#sayfayi_ziyaret_et(driver, 'https://www.sokmarket.com.tr/meyve-ve-sebze-c-20') # Sayfayı ziyaret eder.

try:
    sayfa_basligi = sayfayi_ziyaret_et(driver, url) # Sayfa başlığını alır.
except:
    yaz("Geçersiz URL girdiniz. Varsayılan URL ile devam edilecek.")
    sayfa_basligi = sayfayi_ziyaret_et(driver, 'https://www.sokmarket.com.tr/meyve-ve-sebze-c-20/') 


sayfayi_kaydir(driver, 60, 0.07) # Sayfayı 20 kez kaydırır ve her kaydırma işleminden sonra 1 saniye bekler.
html_verileri = driver.page_source # Tarayıcıdan html verilerini alır.
soup = BeautifulSoup(html_verileri, 'html.parser')  # BeautifulSoup ile html verilerini parse eder.
veri_listesi = verileri_al(soup) # Veri listesini alır.
#verileri_yaz(veri_listesi) # Veri listesini dosyaya yazar.
excele_yaz(veri_listesi, sayfa_basligi) #excel'e yaz verileri ve sayfa başlığını kullan.
driver.quit() #tarayıcıyı kapat